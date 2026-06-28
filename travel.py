# Project State Travel Data Report Group: Duriel H., Leo M., Cynthia B, RoShawn F., Oye O.
import requests
from openpyxl import Workbook
states_url = "https://state-data-api.proxy.beeceptor.com/state-details"
travel_url = "https://state-data-api.proxy.beeceptor.com/state-travel-data"

states_response = requests.get(states_url)
states_data = states_response.json()

travel_response = requests.get(travel_url)
travel_data = travel_response.json()

#combined dataset
combine_data = []

# Loop through states and find the matching travel data

for state_item in states_data:
    current_state_name = state_item["state"]
    
    for travel_item in travel_data:
        if travel_item["state"] == current_state_name:
            
            # Create a combined dictionary with all fields
            
            merge_record = {
                "state": state_item["state"],
                "region": state_item["region"],
                "population": state_item["population"],
                "top_attraction": travel_item["top_attraction"],
                "average_trip_cost": travel_item["average_trip_cost"]
            }
            combine_data.append(merge_record)
            print(merge_record)

# Create Excel Workbook

wb = Workbook()
ws = wb.active
ws.title = "State Travel Report"

# #Write headers

headers = ["State", "Region", "Population", "Top Attraction", "Average Trip Cost"]

for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

#Write combine travel report

for row_num, travel_report in enumerate(combine_data, 2):
    for col_num, value in enumerate(travel_report.values(), 1):
        ws.cell(row=row_num, column=col_num, value=str(value)) 
ws2 = wb.create_sheet("States by Region")

headers2 = ["State", "Top Attraction", "Average Trip Cost", "Region"]

for col_num, header in enumerate(headers2, 1):
    ws2.cell(row=1, column=col_num, value=header)
    
regions = ["Midwest", "West", "South", "Northeast"]

current_row = 2

for region in regions:
    for record in combine_data:
        if record["region"] == region:
            for col_num, value in enumerate([record["state"], record["top_attraction"], record["average_trip_cost"], record["region"]], 1):
                ws2.cell(row=current_row, column=col_num, value=str(value))
            current_row += 1
    current_row += 1
        
   
wb.save("state_travel_report.xlsx")       
    
